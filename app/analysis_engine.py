from __future__ import annotations

from dataclasses import dataclass
from typing import Dict, Tuple, List, Optional, Any

from openpyxl import load_workbook

from app.fin_mapping import map_item_to_key, normalize_text, explain_key

# Eski şema (geriye uyum)
SHEET_BS = "BILANCO"
SHEET_IS = "GELIR"


def _as_float(v: Any) -> float:
    try:
        if v is None:
            return 0.0
        if isinstance(v, (int, float)):
            return float(v)
        s = str(v).strip()
        if s == "":
            return 0.0
        # TR format: 1.234.567,89
        s = s.replace(".", "").replace(",", ".")
        return float(s)
    except Exception:
        return 0.0


# ============================================================
# 1) MİZAN / TRIAL BALANCE PARSER (KODA DAYALI, TASARIM-BAĞIMSIZ)
#    ✅ 3-haneli ana hesap konsolidasyonu + kontra (-) düzeltmesi
# ============================================================

# TDHP kontra ( - ) çalışan 3-haneli hesaplar (minimum kritik liste)
CONTRA_3DIGIT = {
    103, 119, 122, 129, 137, 139, 158, 199,
    222, 224, 229, 237, 239,
    241, 243, 247, 249,
    257, 268, 278, 298, 299,
    302, 308, 402, 408,
    501, 503, 580, 591,
}


def _first3(code_digits: str) -> Optional[int]:
    if not code_digits or len(code_digits) < 3:
        return None
    try:
        return int(code_digits[:3])
    except Exception:
        return None


def _is_contra_name(name: str) -> bool:
    if not name:
        return False
    n = str(name)
    return "(-" in n or "(-)" in n or n.strip().endswith("(-)") or " (-)" in n


@dataclass
class TBRow:
    code: str
    name: str
    balance: float
    code3: int
    digits_len: int


def _looks_like_trial_balance(ws) -> bool:
    for r in range(1, min(ws.max_row, 30) + 1):
        row = [ws.cell(r, c).value for c in range(1, min(ws.max_column, 15) + 1)]
        norm = " ".join(normalize_text(x) for x in row if x is not None)
        if "hesap kodu" in norm and ("bakiye" in norm or "borc" in norm or "alacak" in norm):
            return True
    return False


def _find_tb_header(ws) -> Tuple[int, Dict[str, int]]:
    header_row = None
    headers: Dict[str, int] = {}

    def norm_cell(x: Any) -> str:
        return normalize_text(x)

    for r in range(1, min(ws.max_row, 40) + 1):
        row = [ws.cell(r, c).value for c in range(1, min(ws.max_column, 25) + 1)]
        normed = [norm_cell(x) for x in row]

        if any(x == "hesap kodu" for x in normed) and any(
            x in {"hesap adi", "hesap adı"} or "hesap ad" in x for x in normed
        ):
            header_row = r
            for idx, h in enumerate(normed, start=1):
                if h == "hesap kodu":
                    headers["code"] = idx
                elif h in {"hesap adi", "hesap adı"} or "hesap ad" in h:
                    headers["name"] = idx
                elif h == "bakiye":
                    headers["balance"] = idx
                elif h == "bakiye borc" or h == "bakiye borç":
                    headers["bal_debit"] = idx
                elif h == "bakiye alacak":
                    headers["bal_credit"] = idx
            break

    if not header_row or "code" not in headers:
        raise ValueError("Mizan sheet'inde header bulunamadı (Hesap Kodu...).")

    if "balance" not in headers and ("bal_debit" not in headers or "bal_credit" not in headers):
        raise ValueError("Mizan sheet'inde 'Bakiye' veya 'Bakiye Borç/Alacak' kolonları bulunamadı.")

    if "name" not in headers:
        headers["name"] = headers["code"] + 1

    return header_row, headers


def _parse_trial_balance_sheet(ws) -> List[TBRow]:
    header_row, col = _find_tb_header(ws)
    out: List[TBRow] = []

    for r in range(header_row + 1, ws.max_row + 1):
        raw_code = ws.cell(r, col["code"]).value
        if raw_code is None:
            continue
        code_txt = str(raw_code).strip()
        if code_txt == "":
            continue

        code_digits = "".join(ch for ch in code_txt if ch.isdigit())
        if len(code_digits) < 3:
            continue

        c3 = _first3(code_digits)
        if c3 is None:
            continue

        name = str(ws.cell(r, col["name"]).value or "").strip()

        if "balance" in col:
            bal = _as_float(ws.cell(r, col["balance"]).value)
        else:
            bal_deb = _as_float(ws.cell(r, col["bal_debit"]).value)
            bal_cred = _as_float(ws.cell(r, col["bal_credit"]).value)
            bal = bal_deb - bal_cred

        if abs(bal) < 1e-6:
            continue

        out.append(
            TBRow(
                code=code_digits,
                name=name,
                balance=float(bal),
                code3=int(c3),
                digits_len=len(code_digits),
            )
        )

    return out


def _pick_trial_balance_ws(wb) -> Optional[Any]:
    for nm in wb.sheetnames:
        n = normalize_text(nm)
        if "mizan" in n:
            ws = wb[nm]
            if _looks_like_trial_balance(ws):
                return ws
    for nm in wb.sheetnames:
        ws = wb[nm]
        if _looks_like_trial_balance(ws):
            return ws
    return None


def _consolidate_to_3digit(rows: List[TBRow]) -> Dict[int, float]:
    bucket: Dict[int, List[TBRow]] = {}
    for r in rows:
        bucket.setdefault(r.code3, []).append(r)

    out: Dict[int, float] = {}

    for code3, lst in bucket.items():
        exact3 = [x for x in lst if x.digits_len == 3]
        use_rows = exact3 if exact3 else lst

        total = 0.0
        for x in use_rows:
            v = float(x.balance)
            if (code3 in CONTRA_3DIGIT) or _is_contra_name(x.name):
                v = -abs(v)
            total += v

        out[code3] = total

    return out


def _sum_3range(b3: Dict[int, float], lo: int, hi: int) -> float:
    return sum(v for k, v in b3.items() if lo <= k <= hi)


def _sum_3set(b3: Dict[int, float], codes: List[int]) -> float:
    return sum(float(b3.get(c, 0.0) or 0.0) for c in codes)


def _sum_prefix(rows: List[TBRow], prefixes: List[str]) -> float:
    """
    ✅ NameError fix + Çifte saymayı engeller:
    - Önce 3-haneli konsolidasyon (b3) üzerinden toplar.
    - Prefix "60" gibi 2 haneliyse 600-699 aralığına eşler
    - Prefix "6" gibi 1 haneliyse 600-699 değil; 600-699 için "60" verin.
    - Prefix "780" gibi 3 haneliyse doğrudan b3[780] gibi.
    """
    b3 = _consolidate_to_3digit(rows)

    total = 0.0
    for p in prefixes:
        pd = "".join(ch for ch in str(p) if ch.isdigit())
        if not pd:
            continue

        if len(pd) == 1:
            lo = int(pd) * 100
            hi = lo + 99
            total += _sum_3range(b3, lo, hi)

        elif len(pd) == 2:
            # 60 -> 600-699
            lo = int(pd) * 10
            lo = lo * 10
            hi = lo + 99
            total += _sum_3range(b3, lo, hi)

        else:
            k = int(pd[:3])
            total += float(b3.get(k, 0.0) or 0.0)

    return total


def _trial_balance_to_canonical(rows: List[TBRow]) -> Dict[str, float]:
    bs: Dict[str, float] = {}

    b3 = _consolidate_to_3digit(rows)

    current_assets = _sum_3range(b3, 100, 199)
    non_current_assets = _sum_3range(b3, 200, 299)
    total_assets = current_assets + non_current_assets

    cash = _sum_3set(b3, [100, 101, 102, 108, 103])
    trade_ar = _sum_3range(b3, 120, 129)
    other_ar = _sum_3range(b3, 131, 139)
    inv = _sum_3set(b3, [150, 151, 152, 153, 157, 158, 159, 170, 171, 172, 173, 178, 179])
    prepaid = float(b3.get(180, 0.0) or 0.0)
    other_ca = _sum_3set(b3, [190, 191, 192, 193, 195, 196, 197, 198, 199])

    cl_signed = _sum_3range(b3, 300, 399)
    current_liabilities = abs(cl_signed)

    ll_signed = _sum_3range(b3, 400, 499)
    long_liabilities = abs(ll_signed)

    st_fin_debt = abs(_sum_3set(b3, [300, 301, 303, 304, 305, 306, 309, 302, 308]))
    lt_fin_debt = abs(_sum_3set(b3, [400, 401, 405, 407, 409, 402, 408]))

    trade_payables = abs(_sum_3range(b3, 320, 329))
    tax_liab = abs(_sum_3set(b3, [360, 361, 368, 369, 391, 392]))

    eq_signed = _sum_3range(b3, 500, 599)
    equity = abs(eq_signed)

    # ✅ DSCR için: 303 - Uzun vadeli kredilerin anapara taksitleri vb (12 ay içinde ödenecek)
    current_portion_lt_debt = abs(float(b3.get(303, 0.0) or 0.0))

    bs["cash_and_equivalents"] = cash
    bs["trade_receivables"] = trade_ar
    bs["other_receivables"] = other_ar
    bs["inventories"] = inv
    bs["prepaid_expenses"] = prepaid
    bs["other_current_assets"] = other_ca
    bs["current_assets_total"] = current_assets

    bs["non_current_assets_total"] = non_current_assets
    bs["total_assets"] = total_assets

    bs["short_term_liabilities"] = current_liabilities
    bs["long_term_liabilities"] = long_liabilities

    bs["short_term_fin_debt"] = st_fin_debt
    bs["long_term_fin_debt"] = lt_fin_debt
    bs["trade_payables"] = trade_payables
    bs["tax_liabilities"] = tax_liab

    bs["equity_total"] = equity
    bs["total_liabilities"] = current_liabilities + long_liabilities

    # ✅ DSCR helper
    bs["current_portion_long_term_debt"] = current_portion_lt_debt

    return bs


def _trial_balance_to_income_statement(rows: List[TBRow]) -> Dict[str, float]:
    """
    Mizan’dan P&L üretimi (fallback).

    Tek Düzen (6xx) + 7/A (78x) + 7/B (797) destek:
      - Revenue: 60 (credit -> negative) => revenue positive = -sum(60x)
      - Discounts: 61 (debit) => positive
      - COGS: 62 (debit) => positive
      - Opex: 63 (debit) => positive
      - Other op income: 64 (credit -> negative) => positive
      - Other op expense: 65 (debit) => positive
      - Finance expense: 66 (debit) OR 780/781/782 OR 797
      - Depreciation/Amortization: 796 (7/B) => EBITDA için
    """
    inc: Dict[str, float] = {}

    gross_sales = -_sum_prefix(rows, ["60"])
    discounts = _sum_prefix(rows, ["61"])
    revenue_net = gross_sales - discounts

    cogs = _sum_prefix(rows, ["62"])
    opex = _sum_prefix(rows, ["63"])

    other_op_income = -_sum_prefix(rows, ["64"])
    other_op_exp = _sum_prefix(rows, ["65"])

    fin_exp_66 = _sum_prefix(rows, ["66"])
    fin_exp_7a = _sum_prefix(rows, ["780", "781", "782"])
    fin_exp_7b = _sum_prefix(rows, ["797"])

    finance_expense = fin_exp_66 + fin_exp_7a + fin_exp_7b

    # ✅ Amortisman (7/B)
    dep_amort = _sum_prefix(rows, ["796"])

    gross_profit = revenue_net - cogs
    ebit = gross_profit - opex + other_op_income - other_op_exp
    ebitda = ebit + dep_amort

    inc["revenue"] = revenue_net
    inc["cogs"] = cogs
    inc["gross_profit"] = gross_profit
    inc["opex"] = opex
    inc["ebit"] = ebit
    inc["finance_expense"] = finance_expense
    inc["interest_expense"] = 0.0

    # ✅ EBITDA alanları
    inc["depreciation_amortization"] = dep_amort
    inc["ebitda"] = ebitda

    return inc


# ============================================================
# 1.5) GELİR SHEET ESNEK PARSER (KOD YOK, KALEM ŞARTI YOK)
# ============================================================

def _looks_like_income_sheet(ws) -> bool:
    needles = ["gelir tablosu", "net satis", "satıs", "satis", "hasilat", "satışların maliyeti",
               "satislarin maliyeti", "brut kar", "faiz", "finansman", "favok", "ebit"]
    for r in range(1, min(ws.max_row, 40) + 1):
        row = [ws.cell(r, c).value for c in range(1, min(ws.max_column, 20) + 1)]
        text = " ".join(normalize_text(x) for x in row if x is not None)
        if any(n in text for n in needles):
            return True
    return False


def _find_income_header(ws) -> Optional[Tuple[int, int, int]]:
    """
    returns: (header_row, desc_col, value_col)
    """
    desc_candidates = {
        "kalem", "aciklama", "açıklama", "hesap adi", "hesap adı", "tanim", "tanım",
        "gelir tablosu kalemi", "hesap", "kalem adi", "kalem adı"
    }
    value_candidates = {
        "tutar", "cari donem", "cari dönem", "donem", "dönem", "amount", "current period"
    }

    def norm(x: Any) -> str:
        return normalize_text(x)

    for r in range(1, min(ws.max_row, 80) + 1):
        row_vals = [ws.cell(r, c).value for c in range(1, min(ws.max_column, 40) + 1)]
        normed = [norm(x) for x in row_vals]

        desc_col = None
        for c, h in enumerate(normed, start=1):
            if h in desc_candidates:
                desc_col = c
                break
            if "aciklama" in h or "açıklama" in h or ("hesap" in h and "adi" in h):
                desc_col = c
                break

        if not desc_col:
            continue

        value_col = None
        for c, h in enumerate(normed, start=1):
            if h in value_candidates:
                value_col = c
                break

        # Yıl kolonu (2024/2023 gibi)
        if value_col is None:
            for c, raw in enumerate(row_vals, start=1):
                if isinstance(raw, int) and 1900 <= raw <= 2200:
                    value_col = c
                    break
                if isinstance(raw, float) and 1900 <= int(raw) <= 2200:
                    value_col = c
                    break
                if isinstance(raw, str) and raw.strip().isdigit():
                    iv = int(raw.strip())
                    if 1900 <= iv <= 2200:
                        value_col = c
                        break

        if value_col:
            return (r, desc_col, value_col)

    return None


def _parse_income_sheet_flexible(ws) -> Tuple[Dict[str, float], List[Dict[str, Any]]]:
    """
    KALEM şartı olmadan gelir tablosu okur.
    """
    hdr = _find_income_header(ws)
    if not hdr:
        raise ValueError("Gelir sheet'inde açıklama+tutar header'ı bulunamadı (esnek parser).")

    header_row, desc_col, value_col = hdr

    out: Dict[str, float] = {}
    log: List[Dict[str, Any]] = []

    def is_noise(name: str) -> bool:
        n = normalize_text(name)
        if not n:
            return True
        if n in {"toplam", "genel toplam", "aciklama", "açıklama", "kalem", "tutar"}:
            return True
        if n.startswith("genel toplam") or n.startswith("toplam"):
            return True
        return False

    empty_streak = 0
    for r in range(header_row + 1, ws.max_row + 1):
        raw_name = ws.cell(r, desc_col).value
        raw_val = ws.cell(r, value_col).value

        if raw_name is None and raw_val is None:
            empty_streak += 1
            if empty_streak >= 25:
                break
            continue
        empty_streak = 0

        if raw_name is None:
            continue

        name = str(raw_name).strip()
        if not name or is_noise(name):
            continue

        val = _as_float(raw_val)
        if abs(val) < 1e-6:
            continue

        key = map_item_to_key(name)
        if key:
            out[key] = out.get(key, 0.0) + float(val)

        log.append({
            "raw": name,
            "norm": normalize_text(name),
            "key": key,
            "key_label": explain_key(key) if key else None,
            "value": float(val),
        })

    return out, log


def _merge_income(preferred: Dict[str, float], fallback: Dict[str, float]) -> Dict[str, float]:
    """
    preferred (gelir sheet) varsa onu kullan, yoksa fallback (mizan).
    0 ise override etme.
    """
    merged = dict(fallback or {})
    for k, v in (preferred or {}).items():
        if v is None:
            continue
        vv = float(v)
        if abs(vv) < 1e-6:
            continue
        merged[k] = vv
    return merged


# ============================================================
# 2) ESKİ PARSER (BILANCO/GELIR) - GERİYE UYUMLU KALSIN
# ============================================================

def _find_kalem_headers(ws, max_scan_rows: int = 25) -> List[Tuple[int, int]]:
    headers: List[Tuple[int, int]] = []
    for r in range(1, min(ws.max_row, max_scan_rows) + 1):
        for c in range(1, ws.max_column + 1):
            v = ws.cell(r, c).value
            if isinstance(v, str) and v.strip().upper() == "KALEM":
                headers.append((r, c))
    headers.sort(key=lambda x: (x[0], x[1]))
    return headers


def _year_cols_from_header(ws, header_row: int, kalem_col: int, max_years: int = 10) -> List[Tuple[int, int]]:
    out: List[Tuple[int, int]] = []
    scanned = 0
    c = kalem_col + 1
    while c <= ws.max_column and scanned < 25 and len(out) < max_years:
        v = ws.cell(header_row, c).value
        yr: Optional[int] = None
        if isinstance(v, int) and 1900 <= v <= 2200:
            yr = v
        elif isinstance(v, float) and 1900 <= int(v) <= 2200:
            yr = int(v)
        elif isinstance(v, str):
            vv = v.strip()
            if vv.isdigit():
                iv = int(vv)
                if 1900 <= iv <= 2200:
                    yr = iv
        if yr is not None:
            out.append((yr, c))
        c += 1
        scanned += 1
    out.sort(key=lambda x: x[0])
    return out


def _is_noise_row(name: str) -> bool:
    n = str(name).strip().upper()
    if n in {"AKTIF", "PASIF", "VARLIKLAR", "KAYNAKLAR"}:
        return True
    if n.startswith("GENEL TOPLAM"):
        return True
    if n == "TOPLAM":
        return True
    return False


def _block_rows(ws, start_row: int, end_row: int, kalem_col: int, value_col: int) -> List[Tuple[str, float]]:
    items: List[Tuple[str, float]] = []
    for r in range(start_row, min(end_row, ws.max_row) + 1):
        k = ws.cell(r, kalem_col).value
        if k is None:
            continue
        name = str(k).strip()
        if not name:
            continue
        if _is_noise_row(name):
            continue
        v = ws.cell(r, value_col).value
        items.append((name, _as_float(v)))
    return items


def _items_to_canonical(items: List[Tuple[str, float]]) -> Tuple[Dict[str, float], List[Dict[str, Any]]]:
    out: Dict[str, float] = {}
    log: List[Dict[str, Any]] = []
    for name, val in items:
        n = normalize_text(name)
        key = map_item_to_key(name)
        if key:
            out[key] = out.get(key, 0.0) + float(val)
        log.append({
            "raw": name,
            "norm": n,
            "key": key,
            "key_label": explain_key(key) if key else None,
            "value": float(val),
        })
    return out, log


# ============================================================
# 3) TEK GİRİŞ NOKTASI: parse_financials_xlsx
# ============================================================

def parse_financials_xlsx(xlsx_path: str) -> dict:
    wb = load_workbook(xlsx_path, data_only=True)

    tb_ws = _pick_trial_balance_ws(wb)
    if tb_ws is not None:
        tb_rows = _parse_trial_balance_sheet(tb_ws)
        bs_canon = _trial_balance_to_canonical(tb_rows)

        # ✅ Mizan'dan fallback P&L
        inc_fallback = _trial_balance_to_income_statement(tb_rows)

        # ✅ Gelir sheet varsa: önce esnek parser dene, olmadı legacy KALEM dene
        inc_preferred: Dict[str, float] = {}
        is_log: List[Dict[str, Any]] = []
        is_items: List[Tuple[str, float]] = []
        is_year = None
        income_mode = "trial_balance_only"

        if SHEET_IS in wb.sheetnames:
            ws_is = wb[SHEET_IS]

            # 1) Esnek parser (KALEM şartı yok)
            try:
                if _looks_like_income_sheet(ws_is):
                    inc_preferred, is_log = _parse_income_sheet_flexible(ws_is)
                    if inc_preferred:
                        income_mode = "income_sheet_flexible"
            except Exception:
                inc_preferred, is_log = {}, []

            # 2) Esnek boşsa legacy KALEM parser dene
            if not inc_preferred:
                is_headers = _find_kalem_headers(ws_is)
                if is_headers:
                    hr, kc = is_headers[0]
                    years = _year_cols_from_header(ws_is, hr, kc)
                    is_year, vc = years[-1] if years else (None, kc + 1)
                    next_hr = is_headers[1][0] if len(is_headers) > 1 else ws_is.max_row + 1
                    is_items = _block_rows(ws_is, start_row=hr + 1, end_row=next_hr - 1, kalem_col=kc, value_col=vc)
                    inc_preferred, is_log = _items_to_canonical(is_items)
                    if inc_preferred:
                        income_mode = "income_sheet_legacy_kalem"

        # ✅ Merge: gelir sheet > mizan
        inc = _merge_income(inc_preferred, inc_fallback)

        return {
            "year_bs": None,
            "year_is": is_year,
            "balance_sheet_raw": [(r.code + " " + r.name, r.balance) for r in tb_rows],
            "income_statement_raw": is_items,
            "balance_sheet": bs_canon,
            "income_statement": inc,
            "mapping_log": {
                "mode": "trial_balance",
                "income_mode": income_mode,
                "trial_balance_rows": [{"code": r.code, "name": r.name, "balance": r.balance} for r in tb_rows],
                "income_statement_mapping": is_log,
            },
        }

    # Legacy: BILANCO/GELIR
    if SHEET_BS not in wb.sheetnames or SHEET_IS not in wb.sheetnames:
        raise ValueError("Bu Excel’de mizan bulunamadı; ayrıca BILANCO/GELIR sheet’leri de yok.")

    ws_bs = wb[SHEET_BS]
    ws_is = wb[SHEET_IS]

    bs_headers = _find_kalem_headers(ws_bs)
    if not bs_headers:
        raise ValueError("BILANCO sheet içinde 'KALEM' başlığı bulunamadı.")

    bs_years_found: List[int] = []
    bs_items_all: List[Tuple[str, float]] = []

    for idx, (hr, kc) in enumerate(bs_headers):
        next_hr = bs_headers[idx + 1][0] if idx + 1 < len(bs_headers) else ws_bs.max_row + 1
        end_row = next_hr - 1
        years = _year_cols_from_header(ws_bs, hr, kc)
        if not years:
            continue
        year, vc = years[-1]
        bs_years_found.append(year)
        items = _block_rows(ws_bs, start_row=hr + 1, end_row=end_row, kalem_col=kc, value_col=vc)
        bs_items_all.extend(items)

    if not bs_items_all:
        raise ValueError("BILANCO sheet'inde okunabilir satır bulunamadı.")

    bs_year = max(bs_years_found) if bs_years_found else None

    # GELIR legacy
    is_headers = _find_kalem_headers(ws_is)
    if not is_headers:
        raise ValueError("GELIR sheet içinde 'KALEM' başlığı bulunamadı.")

    hr, kc = is_headers[0]
    years = _year_cols_from_header(ws_is, hr, kc)
    if not years:
        raise ValueError("GELIR sheet'inde yıl kolonları bulunamadı.")
    is_year, vc = years[-1]
    next_hr = is_headers[1][0] if len(is_headers) > 1 else ws_is.max_row + 1
    is_items = _block_rows(ws_is, start_row=hr + 1, end_row=next_hr - 1, kalem_col=kc, value_col=vc)

    bs_canon, bs_log = _items_to_canonical(bs_items_all)
    is_canon, is_log = _items_to_canonical(is_items)

    return {
        "year_bs": bs_year,
        "year_is": is_year,
        "balance_sheet_raw": bs_items_all,
        "income_statement_raw": is_items,
        "balance_sheet": bs_canon,
        "income_statement": is_canon,
        "mapping_log": {
            "mode": "legacy_bs_is",
            "balance_sheet": bs_log,
            "income_statement": is_log,
            "unmapped_balance_sheet": [x for x in bs_log if not x["key"]],
            "unmapped_income_statement": [x for x in is_log if not x["key"]],
        },
    }


# ============================================================
# 4) ANALİZ (ORANLAR) — DOĞRU PAY/PAYDA + EK METRİKLER
# ============================================================

def analyze_financials(fin: dict, sector: str) -> dict:
    bs = fin.get("balance_sheet", {}) or {}
    inc = fin.get("income_statement", {}) or {}

    def g(d: Dict[str, float], key: str) -> float:
        return float(d.get(key, 0.0) or 0.0)

    cash = g(bs, "cash_and_equivalents")
    ar = g(bs, "trade_receivables")
    inv = g(bs, "inventories")

    ca = g(bs, "current_assets_total")
    cl = g(bs, "short_term_liabilities")

    debt_st = g(bs, "short_term_fin_debt")
    debt_lt = g(bs, "long_term_fin_debt")
    cp_lt = g(bs, "current_portion_long_term_debt")  # ✅ DSCR helper (303)
    equity = g(bs, "equity_total")
    total_assets = g(bs, "total_assets")

    revenue = g(inc, "revenue")
    cogs = g(inc, "cogs")
    ebit = g(inc, "ebit")
    fin_exp = g(inc, "finance_expense")
    interest_exp = g(inc, "interest_expense")
    dep_amort = g(inc, "depreciation_amortization")
    ebitda = g(inc, "ebitda")

    if ebitda <= 0:
        # fallback: gelir sheet EBITDA map etmemişse
        ebitda = ebit + dep_amort if dep_amort else ebit

    if ca <= 0:
        ca = cash + ar + inv + g(bs, "other_current_assets") + g(bs, "other_receivables") + g(bs, "prepaid_expenses")
    if cl <= 0:
        cl = g(bs, "trade_payables") + debt_st + g(bs, "tax_liabilities") + g(bs, "provisions_st") + g(bs, "lease_liabilities_st")

    current_ratio = (ca / cl) if cl else None
    quick_ratio = ((ca - inv) / cl) if cl else None

    hard_quick_ratio = ((cash + ar) / cl) if cl else None
    cash_ratio = (cash / cl) if cl else None

    nwc = ca - cl

    net_debt = (debt_st + debt_lt) - cash
    debt_to_equity = ((debt_st + debt_lt) / equity) if equity and equity > 0 else None

    fin_cost = fin_exp if fin_exp else interest_exp
    interest_cover = (ebit / fin_cost) if fin_cost else None
    gross_margin = ((revenue - cogs) / revenue) if revenue else None

    # ✅ DSCR (approx)
    principal_due_12m = (debt_st + cp_lt) if (debt_st or cp_lt) else 0.0
    debt_service = (fin_exp or 0.0) + (principal_due_12m or 0.0)
    dscr = (ebitda / debt_service) if debt_service else None

    bullets: List[str] = []

    if current_ratio is None:
        bullets.append("Cari oran hesaplanamadı (KVYK bulunamadı).")
    elif current_ratio < 1:
        bullets.append(f"Cari oran düşük ({current_ratio:.2f}). KV borç ödeme baskısı riski var.")
    elif current_ratio < 1.5:
        bullets.append(f"Cari oran sınırda ({current_ratio:.2f}). Nakit tamponu güçlendirilebilir.")
    else:
        bullets.append(f"Cari oran iyi ({current_ratio:.2f}). Kısa vadeli likidite daha rahat görünüyor.")

    if quick_ratio is not None:
        if quick_ratio < 1:
            bullets.append(f"Asit-test oranı düşük ({quick_ratio:.2f}). Stoksuz çevrimde zorlanma riski.")
        else:
            bullets.append(f"Asit-test oranı kabul edilebilir ({quick_ratio:.2f}).")

    if hard_quick_ratio is not None:
        bullets.append(f"Sert asit-test (Nakit+Alacak/KVYK): {hard_quick_ratio:.2f}.")
    if cash_ratio is not None:
        bullets.append(f"Nakit oranı (Cash Ratio): {cash_ratio:.2f}.")

    bullets.append(f"Net işletme sermayesi: {nwc:,.0f}.")
    bullets.append(f"Net borç (Finansal borç - nakit): {net_debt:,.0f} (yaklaşık).")

    if debt_to_equity is None:
        bullets.append("Borç/Özkaynak hesaplanamadı (özkaynak 0/negatif veya yok).")
    else:
        bullets.append(f"Borç/Özkaynak: {debt_to_equity:.2f}.")

    if interest_cover is None:
        bullets.append("Faiz karşılama hesaplanamadı (finansman/faiz gideri yok/0).")
    else:
        bullets.append(f"Faiz karşılama: {interest_cover:.2f}.")

    if gross_margin is not None:
        bullets.append(f"Brüt marj yaklaşık %{gross_margin*100:.1f}.")

    # ✅ 11. madde: DSCR
    if dscr is None:
        bullets.append("DSCR hesaplanamadı (borç servisi=0 veya veri eksik).")
    else:
        bullets.append(f"DSCR (yaklaşık): {dscr:.2f} (EBITDA / (Faiz + 12A anapara)).")

    bullets.append("Öneri: 13-hafta nakit projeksiyonu + borç vade haritasını tek sayfada izleyelim.")

    return {
        "meta": {"year_bs": fin.get("year_bs"), "year_is": fin.get("year_is")},
        "metrics": {
            "current_ratio": current_ratio,
            "quick_ratio": quick_ratio,
            "hard_quick_ratio": hard_quick_ratio,
            "cash_ratio": cash_ratio,
            "nwc": nwc,
            "current_assets": ca,
            "current_liabilities": cl,
            "cash": cash,
            "trade_receivables": ar,
            "inventories": inv,
            "net_debt": net_debt,
            "debt_to_equity": debt_to_equity,
            "interest_cover": interest_cover,
            "gross_margin": gross_margin,
            "revenue": revenue,
            "cogs": cogs,
            "equity": equity,
            "total_assets": total_assets,

            # ✅ EBITDA & DSCR metrics
            "depreciation_amortization": dep_amort,
            "ebitda": ebitda,
            "current_portion_long_term_debt": cp_lt,
            "principal_due_12m": principal_due_12m,
            "debt_service": debt_service,
            "dscr": dscr,
        },
        "bullets": bullets[:11],
        "mapping_log": fin.get("mapping_log", {}),
    }
